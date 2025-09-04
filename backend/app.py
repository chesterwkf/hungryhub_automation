import io
import base64
import re
import json

from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from werkzeug.utils import secure_filename
from collections import defaultdict

# Retrieve environment variables
import os
from dotenv import load_dotenv

# Debugging
import logging

# For excel file generation
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# API keys
import google.generativeai as genai
from google.generativeai import types
from google.generativeai.types import GenerationConfig

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
CORS(app)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Access the claude API key
claude_api_key = os.getenv("ANTHROPIC_API_KEY")
gemini_api_key = os.getenv("GEMINI_API_KEY")

# To test the API
@app.route('/api/hello', methods=['GET'])
def hello():
    return jsonify({"message": "Hello from Flask!"})

# Base upload folder
UPLOAD_FOLDER = './uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# Ensure the base upload folder exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    """Check if the file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def encode_image(image_path):
    """
    Encode an image file to base64.

    :param image_path: Path to the image file
    :return: Base64 encoded string of the image
    """
    try:
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except FileNotFoundError:
        logging.error(f"Image file not found: {image_path}")
        return None
    except Exception as e:
        logging.error(f"Error encoding image {image_path}: {e}")
        return None

def extract_text_from_image(api_key, image_paths, prompt_text):
    """
    Use Gemini to extract text from multiple images.

    :param api_key: Your Google API key
    :param image_paths: List of paths to image files
    :param prompt_text: Text prompt for Gemini
    :return: Extracted text from the images
    """
    # Configure the Gemini API
    try:
        genai.configure(api_key=api_key)
    except Exception as e:
        logging.error(f"Error configuring Gemini API: {e}")
        return None

    try:
        # Initialize the Gemini model
        model = genai.GenerativeModel('gemini-1.5-pro-latest')

        # Prepare parts for the request
        parts = []
        image_added = False
        
        # Process each image
        for image_path in image_paths:
            try:
                # Read the image file directly
                with open(image_path, "rb") as image_file:
                    image_data = image_file.read()
                
                # Determine mime type based on file extension
                ext = os.path.splitext(image_path)[1][1:].lower()
                mime_type = "image/jpeg"  # Default
                if ext == "png":
                    mime_type = "image/png"
                elif ext == "gif":
                    mime_type = "image/gif"
                elif ext == "webp":
                    mime_type = "image/webp"
                
                # Add image to parts
                parts.append({
                    "inline_data": {
                        "mime_type": mime_type,
                        "data": base64.b64encode(image_data).decode('utf-8')
                    }
                })
                image_added = True
                logging.info(f"Added image: {os.path.basename(image_path)}")
            except Exception as e:
                logging.warning(f"Error processing image {image_path}: {e}")
        
        if not image_added:
            logging.error("No images could be successfully processed.")
            return None
        
        # Add the text prompt
        parts.append({"text": prompt_text})

        # Send request to Gemini
        logging.info("Sending request to Gemini API...")
        response = model.generate_content(
            parts,
            generation_config=genai.types.GenerationConfig(
                max_output_tokens=4000,
                temperature=0.2,
                top_p=0.8,
            )
        )

        # Extract and return the text response
        if response and hasattr(response, 'text'):
            extracted_text = response.text
            logging.info(f"Received response from Gemini. Content length: {len(extracted_text)} characters.")
            return extracted_text
        else:
            logging.error("Received empty or unexpected response format from Gemini.")
            return None

    except Exception as e:
        logging.error(f"An error occurred during Gemini API call: {e}", exc_info=True)
        return None

def parse_menu_items(extracted_text):
    """
    Parse menu items, prices, and categories from extracted text.

    :param extracted_text: Text extracted from the image by Gemini
    :return: Dictionary of menu items {name: {'price': float, 'category': str}}, plus confidence scores {name: float}
    """
    menu_items = {}
    confidence_scores = {}

    # Basic check if extraction returned anything
    if not extracted_text or not extracted_text.strip():
        logging.warning("Received empty or null text from extraction.")
        return {}, {}

    # Split the text into lines
    lines = extracted_text.strip().split('\n')
    total_non_empty_lines = sum(1 for line in lines if line.strip())
    successful_extractions = 0
    unclear_items = 0
    default_category_items = 0

    logging.info(f"Parsing {total_non_empty_lines} non-empty lines from Gemini output.")

    # --- Refined Regex Patterns ---
    # Price pattern: Handles various currency symbols (optional), commas, and decimals
    # Makes currency symbol and spacing optional. Allows integer prices.
    price_pattern = r'(?:[$€£¥]\s*|\b)(\d+(?:,\d{3})*(?:\.\d{1,2})?)\b'

    for line_num, line in enumerate(lines):
        line = line.strip()
        # Skip empty lines or potential headers/footers from Gemini
        if not line or line.startswith("Here are the items") or line.startswith("---"):
            continue

        line_confidence = 1.0  # Start with full confidence

        # Check for uncertainty indicators in the text (global line check)
        uncertainty_phrases = ["unclear", "can't make out", "illegible", "not visible", "hard to read", "possibly", "maybe", "appears to be"]
        is_unclear = False
        for phrase in uncertainty_phrases:
            if phrase in line.lower():
                line_confidence *= 0.6  # Reduce confidence if uncertainty is indicated globally
                is_unclear = True

        # 1. Find the price first (more reliable anchor)
        price_match = re.search(price_pattern, line)

        if price_match:
            # Extract the price and convert to float
            price_str = price_match.group(1).replace(',', '')
            try:
                price = float(price_str)
            except ValueError:
                logging.warning(f"Line {line_num+1}: Could not convert price '{price_str}' to float. Skipping line: '{line}'")
                continue

            # 2. Extract Item Name (everything before the price match)
            item_name = line[:price_match.start()].strip()
            # Clean up common trailing characters before price
            item_name = re.sub(r'[.…\-_*\s]+$', '', item_name).strip()

            # 3. Extract Category (everything after the price match)
            category_name = line[price_match.end():].strip()
            # Clean up common leading characters after price
            category_name = re.sub(r'^[.…\-_*\s]+', '', category_name).strip()

            # --- Data Cleaning and Validation ---
            # Handle empty item name (likely parsing error or header)
            if not item_name:
                 logging.warning(f"Line {line_num+1}: Extracted empty item name. Skipping line: '{line}'")
                 continue

            # Handle empty or default category
            if not category_name or category_name.lower() in ["unclear", "unknown", "n/a", "none", "-", "--"]:
                category_name = "Uncategorized" # Standardize default
                line_confidence *= 0.8 # Slightly reduce confidence if category was unclear/missing
                default_category_items += 1

            # Remove any "(unclear)" tags added by Gemini from name/category
            item_name = item_name.replace("(unclear)", "").strip()
            category_name = category_name.replace("(unclear)", "").strip()

            # If the global 'unclear' flag was set, mark item count
            if is_unclear:
                unclear_items += 1

            # Check name/category quality (very short strings might be noise)
            if len(item_name) < 3:
                line_confidence *= 0.8
            if len(category_name) < 3 and category_name != "Uncategorized": # Allow short defaults
                 line_confidence *= 0.9 # Less penalty for short category

            # Check price reasonableness
            if price <= 0 or price > 1000: # Adjusted range slightly
                line_confidence *= 0.7

            # Store the item
            successful_extractions += 1
            item_key = item_name # Use name as the key

            # Handle duplicate item names (append index if price/category differ)
            if item_key in menu_items:
                 existing_item = menu_items[item_key]
                 # Only create variant if price OR category is different
                 if existing_item['price'] != price or existing_item['category'] != category_name:
                     count = 1
                     new_key = f"{item_name} ({count})"
                     while new_key in menu_items:
                         count += 1
                         new_key = f"{item_name} ({count})"
                     item_key = new_key
                     logging.info(f"Duplicate item name '{item_name}' found with different price/category. Storing as '{item_key}'.")
                 else:
                     # Exact duplicate, maybe increase confidence slightly? Or just skip. Let's skip.
                     logging.info(f"Exact duplicate item found: '{item_name}'. Skipping.")
                     successful_extractions -= 1 # Decrement success counter as we skipped it
                     continue # Skip to next line

            menu_items[item_key] = {'price': price, 'category': category_name}
            confidence_scores[item_key] = max(0.0, min(1.0, line_confidence)) # Clamp confidence 0-1

        else:
            # Line did not contain a recognizable price pattern
            logging.warning(f"Line {line_num+1}: No valid price pattern found. Skipping line: '{line}'")

    # Calculate overall extraction quality metrics
    extraction_rate = successful_extractions / total_non_empty_lines if total_non_empty_lines > 0 else 0
    avg_confidence = sum(confidence_scores.values()) / len(confidence_scores) if confidence_scores else 0

    # Log the extraction metrics
    logging.info(f"Parsing complete. Successfully extracted {successful_extractions} items.")
    logging.info(f"Extraction rate: {extraction_rate:.1%} of non-empty lines.")
    logging.info(f"Items marked unclear by Gemini or parser: {unclear_items}")
    logging.info(f"Items assigned default 'Uncategorized' category: {default_category_items}")
    logging.info(f"Average confidence score for extracted items: {avg_confidence:.1%}")

    return menu_items, confidence_scores

def categorize_menu_items_with_beverage_priority(menu_items):
    """
    Categorize menu items with enhanced beverage detection and price validation.
    
    :param menu_items: Dictionary {name: {'price': float, 'category': str}}
    :return: Dictionary {category_letter: [{'name': str, 'price': float, 'category': str}, ...]}
    """
    if not menu_items:
        logging.warning("No menu items provided for categorization.")
        return {}
    
    # First, clean up any obviously wrong prices
    cleaned_menu_items = {}
    for name, item_data in menu_items.items():
        price = item_data['price']
        
        # Fix zero or negative prices (invalid data)
        if price <= 0:
            logging.warning(f"Found invalid price for {name}: ${price:.2f}, setting to minimum price of $1.00")
            item_data['price'] = 1.0  # Set a minimum price to avoid division by zero
        
        # Check for unreasonable prices (over 500 baht for a menu item)
        elif price > 500:
            logging.warning(f"Found potentially incorrect price for {name}: ${price:.2f}")
            
            # Extract price from item name if possible
            price_match = re.search(r'฿(\d+)', name)
            if price_match:
                corrected_price = float(price_match.group(1))
                logging.info(f"Corrected price from name for {name}: ${corrected_price:.2f}")
                item_data['price'] = corrected_price
        
        cleaned_menu_items[name] = item_data
    
    # Prepare items list
    items_list = [{'name': name, **data} for name, data in cleaned_menu_items.items()]
    
    # Initialize categories
    alcoholic_beverages = []
    non_alcoholic_beverages = []
    other_items = []
    
    # Process each menu item
    for item in items_list:
        name = item['name'].lower()
        category = item['category'].lower() if item['category'] else ''
        
        # DEFINITIVE RULES FOR NON-ALCOHOLIC BEVERAGES
        if ('pepsi' in name or 
            'coca' in name or 
            'sprite' in name or 
            'fanta' in name or 
            'juice' in name or 
            'soda' in name or 
            'water' in name or 
            'tea' in name or 
            'coffee' in name or 
            'น้ำ' in name or  # Thai for water/drink
            'fizz' in name or 
            'mocktail' in name or 
            'peony' in name or 
            'blue asian' in name or 
            'tropical summer' in name or 
            'yu-ney' in name or 
            'freshy island' in name or 
            'chrysanthemum' in name or 
            'grass jelly' in name):
            
            # EXCLUSIONS: Food items with drink words
            if ('soup' in name or 'dessert' in category.lower() or 
                'ginkgo' in name or 'bua loi' in name or 'rice' in name):
                other_items.append(item)
            else:
                # Ensure correct category
                item['category'] = 'Non-Alcoholic Beverage'
                non_alcoholic_beverages.append(item)
            continue
            
        # SPECIAL CASE: Santa Vittoria
        if 'santa vittoria' in name or 'ซานตา วิตตอเรีย' in name:
            # Santa Vittoria is mineral water (non-alcoholic), whether sparkling or still
            item['category'] = 'Non-Alcoholic Beverage'
            non_alcoholic_beverages.append(item)
            continue
            
        # DEFINITIVE RULES FOR ALCOHOLIC BEVERAGES
        if ('beer' in name or 
            'เบียร์' in name or 
            'singha' in name or 
            'tsingtao' in name or 
            'สิงห์' in name or 
            'ชิงเต่า' in name or 
            'wine' in name or 
            'whisky' in name or 
            'whiskey' in name or 
            'vodka' in name or 
            'gin' in name or 
            'rum' in name or 
            'tequila' in name or 
            'dimple' in name):
                
            # Ensure correct category
            item['category'] = 'Alcoholic Beverage'
            alcoholic_beverages.append(item)
            continue
            
        # Use category information as fallback
        if 'beverage' in category:
            if 'alcoholic' in category:
                alcoholic_beverages.append(item)
            else:
                non_alcoholic_beverages.append(item)
        else:
            # Not a beverage or not clearly identified
            other_items.append(item)
    
    # Create result categories dict
    categories = {}
    
    # Assign alcoholic beverages to category Z
    if alcoholic_beverages:
        categories['Z'] = sorted(alcoholic_beverages, key=lambda x: x['price'], reverse=True)
        
    # Assign non-alcoholic beverages to category Y
    if non_alcoholic_beverages:
        categories['Y'] = sorted(non_alcoholic_beverages, key=lambda x: x['price'], reverse=True)
    
    # Apply dynamic price categorization to non-beverage items
    if other_items:
        # Sort by price (descending)
        sorted_items = sorted(other_items, key=lambda x: x['price'], reverse=True)
        
        # Setup thresholds
        expensive_threshold_ratio = 1.3  # 30% difference
        inexpensive_threshold_ratio = 2.0  # 2x difference
        
        # Price below which we switch threshold
        switch_threshold = sorted_items[0]['price'] / 2 if sorted_items else 0
        
        # Keep track of remaining items to categorize
        remaining_items = sorted_items.copy()
        current_category = 0  # Start with category A (0)
        
        while remaining_items:
            current_category_letter = chr(65 + current_category)  # A, B, C, ...
            category_items = []
            
            # Add the first item to the current category
            first_item = remaining_items.pop(0)
            category_items.append(first_item)
            highest_price = lowest_price = first_item['price']
            
            # Determine which threshold to use
            threshold_ratio = expensive_threshold_ratio if highest_price >= switch_threshold else inexpensive_threshold_ratio
            
            # Add more items within threshold
            while remaining_items:
                next_item = remaining_items[0]
                next_price = next_item['price']
                
                # Prevent division by zero
                if next_price <= 0:
                    logging.warning(f"Item '{next_item['name']}' has invalid price ${next_price:.2f}. Moving to next item.")
                    remaining_items.pop(0)  # Remove this item and continue
                    continue
                
                if highest_price / next_price > threshold_ratio:
                    break
                
                category_items.append(remaining_items.pop(0))
                lowest_price = next_price
            
            # Store the category
            categories[current_category_letter] = category_items
            
            # Prepare for next category
            current_category += 1
            
            # Adjust threshold if needed
            if remaining_items and remaining_items[0]['price'] < switch_threshold:
                threshold_ratio = inexpensive_threshold_ratio
    
    # Log the categorization results
    logging.info("\n--- Final Categorization Results ---")
    for category_letter, items in sorted(categories.items()):
        if not items:
            continue  # Skip empty categories
            
        category_type = "Alcoholic Beverages" if category_letter == 'Z' else \
                       "Non-Alcoholic Beverages" if category_letter == 'Y' else \
                       f"Price Category"
        price_range = f"${min(item['price'] for item in items):.2f} - ${max(item['price'] for item in items):.2f}" if items else "N/A"
        logging.info(f"Category {category_letter} ({category_type}): {len(items)} items. Price range: {price_range}")
        for item in items:
            logging.info(f"  ${item['price']:.2f} - {item['name']}")
    
    return categories


def process_menu_images(api_key, image_folder, num_price_categories=3):
    """
    Process all menu images in a folder, extract items (name, price, category),
    and categorize them by price tiers (A, B, C...).

    :param api_key: Your Gemini API key
    :param image_folder: Path to folder containing menu images
    :param num_price_categories: Number of price categories (A, B, C...)
    :return: Price-categorized menu items dict, and overall confidence scores dict
    """
    # List all image files in the folder
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp'] # Added webp
    image_paths = []

    if not os.path.isdir(image_folder):
        logging.error(f"Image folder not found or is not a directory: {image_folder}")
        return {}, {}

    for file in os.listdir(image_folder):
        if any(file.lower().endswith(ext) for ext in image_extensions):
            image_paths.append(os.path.join(image_folder, file))

    if not image_paths:
        logging.error(f"No supported image files ({', '.join(image_extensions)}) found in {image_folder}")
        return {}, {}

    logging.info(f"Found {len(image_paths)} images to process in folder: {image_folder}")

    # Store all items extracted across all images before final categorization
    # Using the structure: {name: {'price': float, 'category': str}}
    all_menu_items_combined = {}
    all_confidence_scores_combined = {}
    image_extraction_counts = []

    # --- Updated Prompt for Gemini ---
    prompt_text = f"""
    Please extract all menu items from the provided image(s).

For each distinct menu item found, provide the following information on a single line:
1. The complete Item Name in English (if available).
2. The Thai name of the item (if available).
3. If only Thai OR only English name is available, just extract what is present and mark the other blank
4. The exact Price (including currency symbol like $ or ฿ if visible, otherwise just the number).
5. The item's Category (e.g., Appetizer, Main Course, Dessert, Beverage, Side Dish).

Format each item STRICTLY as:
English Item Name Thai Item Name $Price Category

Examples (menu item is available in both english and thai):
Classic Caesar Salad ซีซาร์สลัดคลาสสิค $12.99 Appetizer
Grilled Salmon Fillet ปลาแซลมอนย่าง $24.50 Main Course
New York Cheesecake นิวยอร์กชีสเค้ก $8.00 Dessert
Iced Tea ชาเย็น $3.50 Beverage
French Fries เฟรนช์ฟรายส์ $5.00 Side Dish

Examples (menu item is only available in english):
Pad Thai $12.99 Main Course
Classic Caesar Salad $12.99 Appetizer
Grilled Salmon Fillet $24.50 Main Course
New York Cheesecake $8.00 Dessert
Iced Tea $3.50 Beverage
French Fries $5.00 Side Dish

Examples (menu item is only available in thai):
ผัดไทย $12.99 Main Course
ซีซาร์สลัดคลาสสิค $12.99 Appetizer
ปลาแซลมอนย่าง $24.50 Main Course
นิวยอร์กชีสเค้ก $8.00 Dessert
ชาเย็น $3.50 Beverage
เฟรนช์ฟรายส์ $5.00 Side Dish

IMPORTANT INSTRUCTIONS:
- List EVERY visible menu item, and do not include the menu item details.
- If you cannot clearly determine any part (Price or Category), use the placeholder "(unclear)" for that specific part.
- If the category is totally unknown, use "Uncategorized".
- Ensure the price is extracted accurately.
- Do NOT add any introductory text, explanations, summaries, or formatting like bullet points or markdown.
- Only output the list of items in the specified format, one item per line.
- If an item seems to span multiple lines in the menu image, combine it into a single logical item name if possible.
- Pay attention to sections or headers in the menu image to help determine the category.

CRITICAL BEVERAGE CATEGORIZATION INSTRUCTIONS:
- You MUST correctly identify ALL beverages and categorize them properly:
  * ALL beers (Singha, Tsingtao, Chang, etc.) MUST be labeled as "Alcoholic Beverage"
  * ALL wines, spirits, whiskeys, cocktails MUST be labeled as "Alcoholic Beverage"
  * ALL sodas, juices, teas, coffees, mocktails MUST be labeled as "Non-Alcoholic Beverage"
  * Santa Vittoria Sparkling is alcoholic, regular Santa Vittoria is non-alcoholic water

EXAMPLES OF CORRECT BEVERAGE CATEGORIZATION:
Tsingtao Beer 青岛啤酒 $130.00 Alcoholic Beverage
Singha Beer สิงห์เบียร์ $110.00 Alcoholic Beverage
Pepsi $30.00 Non-Alcoholic Beverage
Orange Juice $80.00 Non-Alcoholic Beverage
Coffee ชาเย็น $75.00 Non-Alcoholic Beverage
Wine ไวน์ $220.00 Alcoholic Beverage
    """

    # Process images (can be done one-by-one or batched if API/model supports it well)
    # Let's stick to one-by-one for simplicity and robustness against single image failures.
    for idx, image_path in enumerate(image_paths):
        logging.info(f"\n--- Processing image {idx+1}/{len(image_paths)}: {os.path.basename(image_path)} ---")

        extracted_text = extract_text_from_image(api_key, [image_path], prompt_text)

        if extracted_text:
            logging.info(f"Text extraction successful for {os.path.basename(image_path)}. Parsing content...")

            # Parse menu items (name, price, category) from this image's text
            menu_items_single, confidence_scores_single = parse_menu_items(extracted_text)
            item_count_single = len(menu_items_single)
            logging.info(f"Parsed {item_count_single} items from {os.path.basename(image_path)}.")
            image_extraction_counts.append((os.path.basename(image_path), item_count_single))

            # Merge items from this image into the combined dictionary
            for item_name, item_data in menu_items_single.items():
                confidence = confidence_scores_single.get(item_name, 0.5) # Default confidence if missing

                # Check if item already exists in combined list
                if item_name in all_menu_items_combined:
                    existing_data = all_menu_items_combined[item_name]
                    # If price or category differs, create a variant
                    if existing_data['price'] != item_data['price'] or existing_data['category'] != item_data['category']:
                        variant_count = 1
                        variant_name = f"{item_name} (Img {idx+1})" # Add image source to variant name
                        while variant_name in all_menu_items_combined:
                             variant_count += 1
                             variant_name = f"{item_name} (Img {idx+1} Var {variant_count})"

                        logging.warning(f"Item '{item_name}' from {os.path.basename(image_path)} differs from previous entry. Storing as '{variant_name}'. "
                                        f"Old: P={existing_data['price']}, C='{existing_data['category']}'. New: P={item_data['price']}, C='{item_data['category']}'.")
                        all_menu_items_combined[variant_name] = item_data
                        all_confidence_scores_combined[variant_name] = confidence
                    else:
                         # Exact same item found again, potentially update confidence if higher?
                         # For now, let's keep the first encountered confidence.
                         logging.info(f"Item '{item_name}' is an exact duplicate from another image. Skipping.")
                else:
                    # New item, add it
                    all_menu_items_combined[item_name] = item_data
                    all_confidence_scores_combined[item_name] = confidence

            logging.info(f"Combined total unique items so far: {len(all_menu_items_combined)}")
        else:
            logging.error(f"Failed to extract text from {os.path.basename(image_path)}. Skipping this image.")
            image_extraction_counts.append((os.path.basename(image_path), 0)) # Record failure

    # --- Post-Processing and Categorization ---
    logging.info(f"\n===== Image Processing Summary =====")
    total_items_extracted = sum(count for _, count in image_extraction_counts)
    logging.info(f"Total items parsed across all images (before deduplication/variants): {total_items_extracted}")
    logging.info(f"Final unique items/variants stored: {len(all_menu_items_combined)}")

    # Print extraction performance by image
    logging.info("\nItems Parsed Per Image:")
    for image_name, item_count in image_extraction_counts:
        logging.info(f"  {image_name}: {item_count} items")

    # Calculate overall average confidence score
    if all_confidence_scores_combined:
        avg_confidence = sum(all_confidence_scores_combined.values()) / len(all_confidence_scores_combined)
        logging.info(f"Overall average extraction confidence: {avg_confidence:.1%}")
    else:
         logging.info("No confidence scores available.")

    if not all_menu_items_combined:
        logging.error("No menu items were successfully extracted from any images.")
        return {}, {}

    # Categorize the combined menu items by price tiers
    logging.info("\n--- Categorizing All Extracted Items by Price ---")
    final_categories = categorize_menu_items_with_beverage_priority(all_menu_items_combined)

    return final_categories, all_confidence_scores_combined

### FOR BUNDLE GENERATION ###
def generate(restaurant_name, average_spend):
    # Configure the API key once
    try:
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            raise ValueError("GEMINI_API_KEY not found in environment variables.")
        genai.configure(api_key=api_key)
    except AttributeError as e:
        print(f"AttributeError during genai.configure: {e}")
        print("This strongly suggests an issue with the 'google-generativeai' library installation or version.")
        print("Please ensure you have the latest version installed (pip install --upgrade google-generativeai) and restart your kernel.")
        exit()
    except Exception as e:
        print(f"Error configuring Gemini API: {e}")
        exit()
    while True:
        try:
            average_spend_per_diner = float(average_spend)
            if average_spend_per_diner <= 0:
                print("Average spend must be a positive number. Please try again.")
                continue
            break
        except ValueError:
            print("Invalid input. Please enter a numeric value for average spend.")
            continue
        except Exception as e:
            print(f"An unexpected error occurred while getting input: {e}")
            return

    # Check if categorized_menu_items.txt exists
    categorized_file_loc = os.path.join(
        app.config['UPLOAD_FOLDER'],
        secure_filename(restaurant_name),
        f"categorized_menu_items_{secure_filename(restaurant_name)}.txt"
    )
    if not os.path.exists(categorized_file_loc):
        print(f"Error: {categorized_file_loc} file not found")

    try:
        # Read the file content
        with open(categorized_file_loc, "r", encoding="utf-8") as file:
            menu_content = file.read()
    except Exception as e:
        print(f"Error reading menu file: {e}")
        return
        
    # Detect categories from menu content
    categories = []
    try:
        # This assumes each line in categorized_menu_items.txt has a category label
        for line in menu_content.split('\n'):
            if line.strip():
                # Extract category from the line (assuming format like "Category X: item")
                parts = line.split(':')
                if len(parts) > 1:
                    category = parts[0].strip()
                    if category.startswith("Category ") and len(category) > 9:
                        cat_letter = category[9:].strip()
                        if cat_letter and cat_letter not in categories:
                            categories.append(cat_letter)
                            
        # Also check for section headers like "----- Category X ($xx.xx - $xx.xx) -----"
        category_pattern = r'----- Category ([A-Z]) \(\$[\d.]+ - \$[\d.]+\) -----'
        section_matches = re.findall(category_pattern, menu_content)
        for match in section_matches:
            if match not in categories:
                categories.append(match)
    except Exception as e:
        print(f"Warning: Error parsing categories: {e}")
        # Fallback to default categories if parsing fails
        categories = ["A", "B", "C", "D"]

    # If no categories were found, use default
    if not categories:
        print("No categories detected. Using default categories A, B, C, D.")
        categories = ["A", "B", "C", "D"]

    # Sort categories alphabetically
    categories.sort()
    
    # Check if beverage categories Y and Z exist
    has_beverages = "Y" in categories or "Z" in categories
    
    # Identify main dish categories (typically A and B)
    main_dish_categories = []
    if "A" in categories:
        main_dish_categories.append("A")
    if "B" in categories:
        main_dish_categories.append("B")
    
    # Identify side dish and appetizer categories (typically C and D)
    side_categories = []
    for cat in categories:
        if cat not in ["Y", "Z"] and cat not in main_dish_categories:
            side_categories.append(cat)
            
    print(f"Detected categories: {categories}")
    if has_beverages:
        print("Beverage categories detected. Will apply special portion rules.")
        
    category_portions_text = "\n".join([f"Category {cat}: [number of items]" for cat in categories])

    # Update prompt with portion rules
    prompt = f"""
Based on these categorised menu data:
{menu_content}

The user has indicated an average spend of ${average_spend_per_diner:.2f} per diner.
Please create 4 UNIQUE menu bundles for a varying number of diners (1-6 diners). Each bundle must include some items from each available menu category ({", ".join(categories)}) where appropriate, following these CRITICAL portion rules:

1. For beverage categories (Y and Z):
   - Each diner should get exactly ONE drink maximum
   - Total beverages should NEVER exceed the number of diners
   - Alcoholic beverages (Z) should not exceed the number of diners
   - Non-alcoholic beverages (Y) should not exceed the number of diners

2. For main dish categories ({", ".join(main_dish_categories)}):
   - Each diner should get 1-2 main dishes maximum
   - For 1-2 diners: 1-3 main dishes total
   - For 3-4 diners: 3-6 main dishes total
   - For 5-6 diners: 5-9 main dishes total

3. For sides, appetizers, and dessert categories ({", ".join(side_categories)}):
   - These are shared items
   - For 1-2 diners: 1-2 items from each category
   - For 3-4 diners: 2-3 items from each category
   - For 5-6 diners: 3-4 items from each category

When creating these bundles, try to ensure the "Price_per_diner" (after discount) is around this average spend of ${average_spend_per_diner:.2f}. It should either be less than 15% or more than 15%.

For each menu bundle, include:
The number of menu items from each category
The number of diners the bundle is designed for
The price per diner (should be around ${average_spend_per_diner:.2f})
A suggested discounted bundle price
Please follow this exact structure for each menu bundle:
Suggested_bundle_price:
Number_of_diners:
category_portions:
{category_portions_text}
Original_bundle_price:
Discount_percentage:
Price_per_diner:
Note: Include all categories in your response, even if some categories might have 0 items in certain bundles.
"""
    model_name = "gemini-2.5-pro-exp-03-25"
    print(f"Using model: {model_name}")

    # Update system instruction with more detailed portion guidelines
    system_instruction_text = f"""You are a menu bundle generator, and your task is to create 4 UNIQUE menu bundles based on a given input of menu categories and a target average spend per diner.

The user has specified an average spend of ${average_spend_per_diner:.2f} per diner. This is a key factor.

Each bundle should be distinctly different from the others and follow these STRICT portion rules:

1. BEVERAGES (Categories Y and Z):
   - Each diner gets exactly ONE drink maximum (never more drinks than diners)
   - Alcoholic beverages (Category Z) should never exceed the number of diners
   - Non-alcoholic beverages (Category Y) should never exceed the number of diners
   - Example: For 2 diners, include at MOST 2 drinks total

2. MAIN DISHES (Categories {", ".join(main_dish_categories)}):
   - Allocate 1-2 main dishes per diner maximum
   - For 1-2 diners: 1-3 main dishes total
   - For 3-4 diners: 3-6 main dishes total
   - For 5-6 diners: 5-9 main dishes total

3. SIDES, APPETIZERS, DESSERTS (Categories {", ".join(side_categories)}):
   - These are usually shared items
   - For 1-2 diners: include 1-2 items from each category
   - For 3-4 diners: include 2-3 items from each category
   - For 5-6 diners: include 3-4 items from each category

For each bundle:
1. Include items from all available categories ({", ".join(["Category " + cat for cat in categories])})
2. Specify exactly how many portions from each category are included
3. Calculate the original price based on the actual menu prices from the input data
4. Apply a reasonable discount (aim for 10-20%)
5. The "Price_per_diner" after the discount should be as close as possible to the user's specified average spend of ${average_spend_per_diner:.2f}

Always include all categories in your response structure, even if some categories have 0 items or don't exist in the input data."""

    try:
        gemini_model = genai.GenerativeModel(
            model_name=model_name,
            system_instruction=system_instruction_text
        )
    except AttributeError as e:
        print(f"AttributeError during genai.GenerativeModel: {e}")
        print("This strongly suggests an issue with the 'google-generativeai' library installation or version.")
        print("Please ensure you have the latest version installed (pip install --upgrade google-generativeai) and restart your kernel.")
        return
    except Exception as e:
        print(f"Error initializing GenerativeModel: {e}")
        return

    contents = [
    {
        "role": "user", 
        "parts": [{"text": prompt}]
    }
    ]

    generation_config_for_method = GenerationConfig(
        response_mime_type="application/json",
    )

    complete_response = ""
    print("\nGenerating menu bundles. This may take a moment...")
    try:
        response_stream = gemini_model.generate_content(
            contents=contents,
            generation_config=generation_config_for_method,
            stream=True
        )

        if response_stream is None:
            print("Warning: API returned None for response stream. This is unexpected with streaming.")
            # Fallback logic (as you had)
            # ...
        else:
            for chunk in response_stream:
                chunk_text = ""
                if hasattr(chunk, 'text') and chunk.text is not None:
                    chunk_text = chunk.text
                elif hasattr(chunk, 'parts') and chunk.parts:
                    for part in chunk.parts:
                        if hasattr(part, 'text') and part.text is not None:
                            chunk_text += part.text
                elif hasattr(chunk, 'candidates') and chunk.candidates:
                    for candidate in chunk.candidates:
                        if hasattr(candidate, 'content') and candidate.content and hasattr(candidate.content, 'parts'):
                            for part in candidate.content.parts:
                                if hasattr(part, 'text') and part.text is not None:
                                    chunk_text += part.text
                
                complete_response += chunk_text
                if chunk_text:
                    print(chunk_text, end="", flush=True)
                else:
                    print(".", end="", flush=True)
            print()

    except Exception as e:
        print(f"\n\nError during generation: {e}")
        # Fallback logic (as you had)
        if not complete_response:
            try:
                print("Attempting to use non-streaming API as final fallback...")
                response = gemini_model.generate_content(
                    contents=contents,
                    generation_config=generation_config_for_method,
                )
                # ... (rest of your fallback logic) ...
                if hasattr(response, 'text') and response.text:
                    complete_response = response.text
                    print(f"\nReceived {len(complete_response)} characters from fallback request")
                elif response and hasattr(response, 'parts'):
                    for part_ in response.parts: # renamed to avoid conflict
                        if hasattr(part_, 'text') and part_.text is not None:
                            complete_response += part_.text
                    print(f"\nReceived {len(complete_response)} characters from fallback request (from parts)")
                else:
                    print("\nFallback response was empty or in an unexpected format.")
            except Exception as fallback_error:
                print(f"Fallback request also failed: {fallback_error}")


    # ... (rest of your JSON parsing and saving code) ...
    # Save the raw response for debugging
    try:
        menu_bundles_raw_file = os.path.join(
            app.config['UPLOAD_FOLDER'],
            secure_filename(restaurant_name),
            f"menu_bundles_raw_{secure_filename(restaurant_name)}.txt"
        )
        with open(menu_bundles_raw_file, "w", encoding="utf-8") as f:        
            f.write(complete_response)
        print(f"\nRaw response saved to menu_bundles_raw.txt ({len(complete_response)} chars)")
    except Exception as e:
        print(f"Error saving raw response: {e}")

    # Parse and save as JSON (your existing logic from here is likely fine)
    try:
        print("\nParsing response as JSON...")
        if not complete_response.strip():
            raise ValueError("Empty response received from API")

        all_valid_jsons = []
        unique_bundles = set()

        cleaned_response = complete_response.strip()
        if cleaned_response.startswith("```json"):
            cleaned_response = cleaned_response[7:]
        if cleaned_response.startswith("```"): # Handle just ``` without json
             cleaned_response = cleaned_response[3:]
        if cleaned_response.endswith("```"):
            cleaned_response = cleaned_response[:-3]
        cleaned_response = cleaned_response.strip()


        try:
            # First, try parsing the (cleaned) entire response as a single JSON object (likely a list of bundles)
            json_data = json.loads(cleaned_response)
            if isinstance(json_data, list):
                all_valid_jsons = json_data[:4]
            elif isinstance(json_data, dict) and "bundles" in json_data and isinstance(json_data["bundles"], list):
                all_valid_jsons = json_data["bundles"][:4]
            elif isinstance(json_data, dict): # If it's a single bundle object
                 all_valid_jsons = [json_data]
            else:
                print("Parsed JSON is not in expected list or dict format, proceeding to regex.")
                raise json.JSONDecodeError("Not a list or expected dict", cleaned_response, 0) # Force fallback

        except json.JSONDecodeError as e:
            print(f"Direct JSON parsing failed: {e}. Trying extraction methods on raw response...")
            
            # Use the original complete_response for regex matching if direct parsing fails
            json_pattern = r'(\{(?:[^{}]|(?:\{(?:[^{}]|(?:\{[^{}]*\}))*\}))*\})' # More robust pattern
            json_matches = re.findall(json_pattern, complete_response)

            if json_matches:
                potential_jsons = sorted(list(set(json_matches)), key=len, reverse=True) # Unique matches
                
                for potential_json_str in potential_jsons:
                    if len(all_valid_jsons) >= 4:
                        break
                    try:
                        # Clean up potential JSON string further if needed
                        # (e.g., remove leading/trailing non-JSON characters if pattern is too greedy)
                        clean_potential_json = potential_json_str.strip()
                        if not (clean_potential_json.startswith('{') and clean_potential_json.endswith('}')):
                            continue # Skip if not a valid JSON object structure

                        json_data = json.loads(clean_potential_json)
                        json_str_for_uniqueness = json.dumps(json_data, sort_keys=True)
                        if json_str_for_uniqueness not in unique_bundles:
                            unique_bundles.add(json_str_for_uniqueness)
                            all_valid_jsons.append(json_data)
                    except json.JSONDecodeError:
                        continue
            
            # Also try code blocks from the original complete_response
            if len(all_valid_jsons) < 4:
                code_block_matches = re.findall(r'```(?:json)?\s*([\s\S]*?)\s*```', complete_response, re.DOTALL)
                for code_block in code_block_matches:
                    if len(all_valid_jsons) >= 4:
                        break
                    try:
                        # If the code block itself contains a list of bundles
                        json_data_block = json.loads(code_block.strip())
                        if isinstance(json_data_block, list):
                            for item in json_data_block:
                                if len(all_valid_jsons) >= 4:
                                    break
                                json_str_for_uniqueness = json.dumps(item, sort_keys=True)
                                if json_str_for_uniqueness not in unique_bundles:
                                    unique_bundles.add(json_str_for_uniqueness)
                                    all_valid_jsons.append(item)
                        elif isinstance(json_data_block, dict): # A single bundle
                            json_str_for_uniqueness = json.dumps(json_data_block, sort_keys=True)
                            if json_str_for_uniqueness not in unique_bundles:
                                unique_bundles.add(json_str_for_uniqueness)
                                all_valid_jsons.append(json_data_block)
                    except json.JSONDecodeError:
                        # If parsing the whole block fails, try finding individual JSON objects within it
                        inner_json_matches = re.findall(json_pattern, code_block)
                        for inner_match_str in inner_json_matches:
                            if len(all_valid_jsons) >= 4:
                                break
                            try:
                                inner_json_data = json.loads(inner_match_str.strip())
                                json_str_for_uniqueness = json.dumps(inner_json_data, sort_keys=True)
                                if json_str_for_uniqueness not in unique_bundles:
                                    unique_bundles.add(json_str_for_uniqueness)
                                    all_valid_jsons.append(inner_json_data)
                            except json.JSONDecodeError:
                                continue

        # Ensure we have only 4 bundles maximum from all combined efforts
        if len(all_valid_jsons) > 4:
            all_valid_jsons = all_valid_jsons[:4]

        if all_valid_jsons:
            # Validate structure of each bundle (optional but good)
            final_bundles = []
            for bundle in all_valid_jsons:
                if isinstance(bundle, dict) and "Suggested_bundle_price" in bundle and "Number_of_diners" in bundle and "category_portions" in bundle:
                    # Add basic portion validation
                    diners = bundle.get("Number_of_diners", 0)
                    portions = bundle.get("category_portions", {})
                    
                    # Check beverage portions if applicable
                    valid_portions = True
                    if has_beverages:
                        total_beverages = 0
                        for bev_cat in ["Y", "Z"]:
                            if bev_cat in portions:
                                bev_count = int(portions[bev_cat]) if isinstance(portions[bev_cat], (int, str)) else 0
                                if bev_count > diners:
                                    print(f"Warning: Bundle has {bev_count} beverages from category {bev_cat} for {diners} diners")
                                    valid_portions = False
                                total_beverages += bev_count
                        
                        if total_beverages > diners:
                            print(f"Warning: Bundle has {total_beverages} total beverages for {diners} diners")
                            valid_portions = False
                    
                    # We'll include it anyway but log warnings
                    if valid_portions:
                        print(f"Valid bundle created for {diners} diners")
                    final_bundles.append(bundle)
                else:
                    print(f"Warning: A parsed JSON object does not match expected bundle structure: {bundle}")
            
            if final_bundles:
                menu_bundles_file = os.path.join(
                    app.config['UPLOAD_FOLDER'],
                    secure_filename(restaurant_name),
                    f"menu_bundles_{secure_filename(restaurant_name)}.json"
                )
                with open(menu_bundles_file, "w", encoding="utf-8") as json_file:
                    json.dump(all_valid_jsons, json_file, indent=4)
                print(f"Saved {len(all_valid_jsons)} unique menu bundles to {menu_bundles_file}")
            else:
                print("No valid JSONs found in the response")
        else:
            print("No valid JSONs found in the response after all parsing attempts.")
            print("Please check menu_bundles_raw.txt to see the actual response.")

    except ValueError as ve: # For the "Empty response" case
        print(f"Error: {ve}")
        print("Please check menu_bundles_raw.txt.")
    except Exception as e:
        print(f"Error: Failed to parse response as JSON: {e}")
        print("Please check menu_bundles_raw.txt to see the actual response.")

def dump_file_contents(file_path, max_lines=20):
    """
    Dump the first few lines of a file to the log for debugging.
    
    Args:
        file_path (str): Path to the file to dump
        max_lines (int): Maximum number of lines to dump
    """
    try:
        absolute_path = os.path.abspath(file_path)
        if not os.path.exists(absolute_path):
            logging.error(f"File not found for debugging: {absolute_path}")
            return
            
        logging.info(f"Dumping first {max_lines} lines of {file_path}:")
        with open(absolute_path, 'r', encoding='utf-8') as f:
            for i, line in enumerate(f):
                if i >= max_lines:
                    logging.info("... (more lines)")
                    break
                logging.info(f"  Line {i+1}: {line.rstrip()}")
    except Exception as e:
        logging.error(f"Error dumping file contents: {e}")

def parse_menu_items_for_excel(file_path='./categorized_menu_items.txt'):
    """
    Parse the categorized menu items text file, expecting format:
    $Price - English Name Thai Name (Category)
    or fallback to:
    $Price - English Name (Category)

    Args:
        file_path (str): Path to the text file containing categorized menu items

    Returns:
        dict: Dictionary with price categories ('Category A', etc.) as keys
            and lists of menu item dicts
            ({'name_en': str, 'name_th': str, 'price': float, 'category': str}) as values.
    """
    # Initialize with basic categories but we'll dynamically add more if needed
    menu_items_by_price_band = {
        "Category A": [],
        "Category B": [],
        "Category C": [],
        "Category D": []
    }

    absolute_path = os.path.abspath(file_path)
    logging.info(f"Attempting to read menu items from: {absolute_path}")

    try:
        if not os.path.exists(absolute_path):
            raise FileNotFoundError(f"File not found at {absolute_path}")

        current_price_band = None # e.g., "Category A", "Category B"
        items_parsed_count = 0

        # Regex V2 (Simplified): Capture combined Name part and Category
        # Group 1: Combined English [Thai] Name part (non-greedy)
        # Group 2: Semantic Category (inside parentheses)
        name_category_pattern = re.compile(r'^(.*?)\s*\(([^)]+)\)$')

        with open(absolute_path, 'r', encoding='utf-8') as file:
            for line_num, line in enumerate(file, 1):
                line = line.strip()

                # Skip empty lines or separator/header lines
                if not line or line.startswith('====='):
                    continue

                # Check if this is a PRICE BAND header line (e.g., ----- Category A -----)
                if line.startswith('----- Category'):
                    try:
                        # Extract price band character (A, B, C, D or any letter)
                        band_match = re.search(r'Category\s+([A-Za-z])', line)
                        if band_match:
                            band_char = band_match.group(1)
                            current_price_band = f"Category {band_char}"
                            
                            # Add this category to our dictionary if it doesn't exist
                            if current_price_band not in menu_items_by_price_band:
                                menu_items_by_price_band[current_price_band] = []
                                logging.info(f"Line {line_num}: Added new price band '{current_price_band}' to dictionary")
                            
                            logging.info(f"Line {line_num}: Found Price Band Header '{current_price_band}'")
                        else:
                            logging.warning(f"Line {line_num}: Could not extract category letter from '{line}'. Ignoring section.")
                            current_price_band = None
                    except IndexError:
                        logging.warning(f"Line {line_num}: Malformed price band header: '{line}'. Ignoring.")
                        current_price_band = None

                # Check if this is a menu item line (starts with '$') and we are inside a valid price band
                elif line.startswith('$') and current_price_band:
                    try:
                        parts = line.split(' - ', 1)
                        if len(parts) == 2:
                            price_str = parts[0].strip()
                            name_and_category_str = parts[1].strip()
                            
                            # Debug: Print the extracted name+category string
                            logging.info(f"Line {line_num}: Extracted name+category: '{name_and_category_str}'")

                            price_value = float(price_str.replace('$', '').replace(',', ''))

                            # Now, parse the "Combined Name (Category)" part
                            match = name_category_pattern.match(name_details_str)

                            item_name_en = "N/A"
                            item_name_th = "" # Default to empty string
                            semantic_category = "Uncategorized" # Default category

                            if match:
                                item_name = match.group(1).strip() # Extract name
                                semantic_category = match.group(2).strip() # Extract category from parentheses
                                
                                # Debug: Print the matched name and category
                                logging.info(f"Line {line_num}: Matched name: '{item_name}', category: '{semantic_category}'")
                            else:
                                # If format doesn't match "Name (Category)", treat whole string as name
                                # and assign a default category. Log a warning.
                                logging.warning(f"Line {line_num}: Item format mismatch. Expected '$Price - Name (Category)', got '$Price - {name_and_category_str}'. Using full string as name.")
                                item_name = name_and_category_str
                                semantic_category = "Uncategorized" # Default semantic category
                                
                                # Debug: Print the fallback name assignment
                                logging.info(f"Line {line_num}: Fallback name: '{item_name}', category: '{semantic_category}'")

                            # Create and append the item data (including semantic category) to the current price band list
                            menu_item = {
                                'name': item_name,
                                'price': price_value,
                                'category': semantic_category,  # Store the extracted semantic category
                                'price_band': current_price_band  # Store which price band this belongs to
                            }
                            
                            # Debug: Print the full item dictionary before appending
                            logging.info(f"Line {line_num}: Adding item: {menu_item}")
                            
                            menu_items_by_price_band[current_price_band].append(menu_item)
                            items_parsed_count += 1
                        else:
                            logging.warning(f"Line {line_num}: Malformed item line. Expected format '$Price - Name (Category)', got: '{line}'.")
                    except ValueError as ve:
                        logging.warning(f"Line {line_num}: Could not parse price from '{parts[0] if len(parts) > 0 else line}'. Error: {ve}. Skipping item: '{line}'.")
                    except Exception as e:
                        logging.error(f"Line {line_num}: Error processing item line: '{line}'. Error: {e}", exc_info=True)
                elif current_price_band and line:
                    logging.info(f"Line {line_num}: Skipping non-item line within {current_price_band}: '{line}'")

        # Debug: Print the final data structure
        logging.info("Final parsed menu items structure:")
        for band, items in menu_items_by_price_band.items():
            logging.info(f"  {band}: {len(items)} items")
            for item in items:
                logging.info(f"    - Name: '{item.get('name', 'MISSING')}', "
                           f"Price: {item.get('price', 'MISSING')}, "
                           f"Category: '{item.get('category', 'MISSING')}'")

        if items_parsed_count > 0:
            logging.info(f"Successfully parsed {items_parsed_count} menu items from {file_path}")
        else:
            logging.warning(f"No menu items were successfully parsed from {file_path}. Check file format or content.")

    except FileNotFoundError as e:
        logging.error(f"{e}")
        logging.warning("Using default menu items because the file could not be read.")
        # Defaults remain the same structure internally
        menu_items_by_price_band["Category A"] = [
            {"name": "Default BBQ ribs", "price": 700.0, "category": "Main Course", "price_band": "Category A"},
            {"name": "Default Chimichurri steak", "price": 590.0, "category": "Main Course", "price_band": "Category A"},
        ]
        menu_items_by_price_band["Category B"] = [
            {"name": "Default Milk Shakes", "price": 190.0, "category": "Beverage", "price_band": "Category B"},
        ]
        menu_items_by_price_band["Category C"] = [
            {"name": "Default Strawberry blast", "price": 100.0, "category": "Beverage", "price_band": "Category C"},
        ]
        menu_items_by_price_band["Category D"] = [
            {"name": "Default Extra shot", "price": 25.0, "category": "Add-on", "price_band": "Category D"},
        ]
        menu_items_by_price_band["Category B"] = [{"name_en": "Default Milk Shakes", "name_th": "มิลค์เชค (ค่าเริ่มต้น)", "price": 190.0, "category": "Beverage"}]
        menu_items_by_price_band["Category C"] = [{"name_en": "Default Strawberry blast", "name_th": "สตรอเบอร์รี่ บลาสท์ (ค่าเริ่มต้น)", "price": 100.0, "category": "Beverage"}]
        menu_items_by_price_band["Category D"] = [{"name_en": "Default Extra shot", "name_th": "เพิ่มช็อต (ค่าเริ่มต้น)", "price": 25.0, "category": "Add-on"}]
    except Exception as e:
        logging.error(f"An unexpected error occurred while reading menu items file: {e}", exc_info=True)
        logging.warning("Using default menu items due to the error.")
        # Populate with defaults including category key
        menu_items_by_price_band["Category A"] = [{"name": "Error Default A", "price": 1.0, "category": "Error", "price_band": "Category A"}]
        menu_items_by_price_band["Category B"] = [{"name": "Error Default B", "price": 1.0, "category": "Error", "price_band": "Category B"}]
        menu_items_by_price_band["Category C"] = [{"name": "Error Default C", "price": 1.0, "category": "Error", "price_band": "Category C"}]
        menu_items_by_price_band["Category D"] = [{"name": "Error Default D", "price": 1.0, "category": "Error", "price_band": "Category D"}]

    # Print summary of parsed items
    logging.info("--- Parsed Menu Items Summary (by Price Band) ---")
    for price_band, items in menu_items_by_price_band.items():
        logging.info(f"  {price_band}: {len(items)} items")
        # Optional: Print first few items for verification
        if items:
            for item in items[:min(3, len(items))]:
                logging.info(f"    - {item.get('name', 'MISSING')} (${item.get('price', 0):.2f}) ({item.get('category', 'MISSING')})")
            if len(items) > 3:
                logging.info("    - ...")
    logging.info("-------------------------------------------------")

    return menu_items_by_price_band # Return the dictionary structured by price bands

def create_raw_menu_sheet(workbook, menu_items_by_price_band):
    """
    Creates a 'Raw Menu Items' sheet displaying all menu items in a flat list format.
    This sheet allows editing and reflects changes in the Menu Items sheet.
    
    Args:
        workbook: The openpyxl workbook to modify
        menu_items_by_price_band: Dictionary structured by price bands, containing lists of item dictionaries
        
    Returns:
        tuple: (sheet, menu_item_row_map) where menu_item_row_map maps items to their row numbers in this sheet
    """
    raw_menu_sheet = workbook.create_sheet(title='Raw Menu Items')
    
    # --- Define styles ---
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # --- Sheet Header ---
    raw_menu_sheet.merge_cells('A1:C1')  # Changed from A1:D1 to A1:C1 (removed Price Band column)
    header_cell = raw_menu_sheet['A1']
    header_cell.value = "Raw Menu Items (All Items Flattened)"
    header_cell.fill = red_fill
    header_cell.font = white_font
    header_cell.alignment = center_aligned
    raw_menu_sheet.row_dimensions[1].height = 25
    
    # --- Column Headers ---
    column_headers = ["Item Name", "Price", "Semantic Category"]  # Removed "Price Band"
    for col, header in enumerate(column_headers, start=1):
        cell = raw_menu_sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center_aligned
        cell.fill = light_gray_fill
        cell.border = border
    
    raw_menu_sheet.row_dimensions[2].height = 20
    
    # --- Set column widths ---
    raw_menu_sheet.column_dimensions['A'].width = 45  # Item Name - made wider
    raw_menu_sheet.column_dimensions['B'].width = 15  # Price
    raw_menu_sheet.column_dimensions['C'].width = 35  # Semantic Category - made wider
    
    # --- Add help text to explain how to use the sheet ---
    help_row = 3
    raw_menu_sheet.merge_cells(f'A{help_row}:C{help_row}')
    help_cell = raw_menu_sheet.cell(row=help_row, column=1)
    help_cell.value = "Edit values in this sheet to automatically update the Menu Items sheet"
    help_cell.font = Font(italic=True, color="FF0000")
    help_cell.alignment = center_aligned
    
    # Starting row for menu items
    current_row = 4
    total_items = 0
    
    # Dictionary to track which row each menu item is in (for referencing from Menu Items sheet)
    menu_item_row_map = {}
    
    # Create a flattened list of all items with their price band
    flattened_items = []
    for price_band, items in menu_items_by_price_band.items():
        for item in items:
            item_with_band = item.copy()  
            item_with_band['price_band'] = price_band  # Include the price band info
            flattened_items.append(item_with_band)
    
    # Sort the flattened items by name for easier reference
    flattened_items.sort(key=lambda x: x.get('name', '').lower())
    
    # Populate the sheet with the sorted items
    for item in flattened_items:
        # Store the row number for this item (keyed by name and price band)
        item_key = (item.get('name', ''), item.get('price_band', ''))
        menu_item_row_map[item_key] = current_row
                
        # Column A: Item Name
        cell = raw_menu_sheet.cell(row=current_row, column=1)
        cell.value = item.get('name', 'N/A')
        cell.alignment = left_aligned
        cell.border = border
        
        # Column B: Price
        cell = raw_menu_sheet.cell(row=current_row, column=2)
        cell.value = item.get('price', 0)
        cell.number_format = '#,##0.00'
        cell.alignment = center_aligned
        cell.border = border
        
        # Column C: Semantic Category
        cell = raw_menu_sheet.cell(row=current_row, column=3)
        cell.value = item.get('category', 'Unknown')
        cell.alignment = left_aligned
        cell.border = border
        
        # Apply alternating row colors
        if current_row % 2 == 0:
            for col in range(1, 4):  # Changed from 1-5 to 1-4 (removed Price Band column)
                raw_menu_sheet.cell(row=current_row, column=col).fill = PatternFill(
                    start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
                )
        
        current_row += 1
        total_items += 1
    
    # --- Add a count summary at the bottom ---
    summary_row = current_row + 1
    raw_menu_sheet.merge_cells(f'A{summary_row}:C{summary_row}')  # Changed from A:D to A:C
    summary_cell = raw_menu_sheet.cell(row=summary_row, column=1)
    summary_cell.value = f"Total Items: {total_items}"
    summary_cell.font = bold_font
    summary_cell.alignment = left_aligned
    
    logging.info(f"'Raw Menu Items' sheet created with {total_items} menu items.")
    return raw_menu_sheet, menu_item_row_map

def create_menu_sheet(workbook, menu_items_by_price_band, menu_item_row_map):
    """
    Creates the 'Menu Items' sheet in the workbook with references to the Raw Menu Items sheet.
    Column A ('Category') now uses the semantic category parsed from the file.
    Column B is English Name, Column C is Thai Name.

    Args:
        workbook: The openpyxl workbook to modify
        menu_items_by_price_band: Dictionary structured by price bands ('Category A', 'B', etc.),
                                containing lists of item dictionaries
        menu_item_row_map: Dictionary mapping item names to their row numbers in the Raw Menu Items sheet
    """
    menu_sheet = workbook.create_sheet(title='Menu Items')

    # --- Styles (same as before) ---
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFFD4", end_color="FFFFD4", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Sheet Headers and Setup ---
    menu_sheet.merge_cells('A1:F1') # Merged across 6 columns
    header_cell = menu_sheet['A1']
    header_cell.value = "Hungry Hub Menu Sections"
    header_cell.fill = red_fill; header_cell.font = white_font; header_cell.alignment = center_aligned
    menu_sheet.row_dimensions[1].height = 25

    menu_sheet['H1'] = "Formula to Calculate NET Price for Menu"
    menu_sheet['H3'] = "VAT (Extra)"; menu_sheet['J3'] = "7%"
    menu_sheet['H4'] = "Service (Extra)"; menu_sheet['J4'] = "10%"
    menu_sheet['H6'] = "You can Adjust"

    column_headers = [
        "Category",             # Col A
        "Menu Name (English)",  # Col B
        "Menu Name (Thai)",     # Col C
        "Description (Optional)",# Col D
        "Menu Price",           # Col E
        "Price (NET) (Formula)" # Col F
    ]
    for col, header in enumerate(column_headers, start=1):
        cell = menu_sheet.cell(row=2, column=col)
        cell.value = header; cell.font = bold_font; cell.alignment = center_aligned; cell.border = border
    menu_sheet.row_dimensions[2].height = 20

    menu_sheet.column_dimensions['A'].width = 25 # Category
    menu_sheet.column_dimensions['B'].width = 40 # Name (English)
    menu_sheet.column_dimensions['C'].width = 40 # Name (Thai)
    menu_sheet.column_dimensions['D'].width = 40 # Description
    menu_sheet.column_dimensions['E'].width = 15 # Menu Price
    menu_sheet.column_dimensions['F'].width = 15 # Price (NET)

    current_row = 3

    # --- Populate Menu Items ---
    # Iterate through all price bands in the dictionary
    for price_band_key in menu_items_by_price_band.keys():
        # Add the Price Band header row (e.g., "Category A")
        menu_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        top_left_cell = menu_sheet.cell(row=current_row, column=1, value=price_band_key)
        top_left_cell.fill = red_fill; top_left_cell.font = white_font; top_left_cell.alignment = center_aligned
        menu_sheet.row_dimensions[current_row].height = 22
        current_row += 1

        items_in_band = menu_items_by_price_band.get(price_band_key, [])

        if not items_in_band:
            logging.info(f"No items found for Price Band {price_band_key} to populate in the 'Menu Items' sheet.")
            continue # Skip to the next price band

        # Add each menu item from the list for this band
        for item_dict in items_in_band:
            # Find this item in the Raw Menu Items sheet
            item_key = (item_dict.get('name', ''), price_band_key)
            if item_key in menu_item_row_map:
                raw_menu_row = menu_item_row_map[item_key]
                
                # Column A: Semantic Category - Link to Raw Menu Items sheet
                cell = menu_sheet.cell(row=current_row, column=1)
                cell.value = f"='Raw Menu Items'!C{raw_menu_row}"  # Reference the Category in Raw Menu Items
                cell.alignment = left_aligned
                cell.border = border

                # Column B: Menu Name - Link to Raw Menu Items sheet
                cell = menu_sheet.cell(row=current_row, column=2)
                cell.value = f"='Raw Menu Items'!A{raw_menu_row}"  # Reference the Name in Raw Menu Items
                cell.alignment = left_aligned
                cell.border = border

                # Column C: Description (Thai) - Placeholder (remains unchanged)
                cell = menu_sheet.cell(row=current_row, column=3, value="")
                cell.alignment = left_aligned
                cell.border = border

                # Column D: Menu Price - Link to Raw Menu Items sheet
                cell = menu_sheet.cell(row=current_row, column=4)
                cell.value = f"='Raw Menu Items'!B{raw_menu_row}"  # Reference the Price in Raw Menu Items
                cell.number_format = '#,##0.00'; cell.alignment = center_aligned; cell.border = border

                # Column E: Price (NET) - Calculated based on the referenced price
                cell = menu_sheet.cell(row=current_row, column=5)
                cell.value = f"=ROUND(D{current_row}*1.177,0)"  # Calculate NET price as a formula
                cell.number_format = '#,##0'; cell.alignment = center_aligned; cell.border = border
            else:
                # Fallback if item not found in Raw Menu Items (shouldn't happen)
                logging.warning(f"Item {item_key} not found in Raw Menu Items sheet. Using direct values.")
                
                # Column A: Category (direct value)
                cell = menu_sheet.cell(row=current_row, column=1, value=item_dict.get('category', 'Unknown'))
                cell.alignment = left_aligned
                cell.border = border
                
                # Column B: Name (direct value)
                cell = menu_sheet.cell(row=current_row, column=2, value=item_dict.get('name', 'N/A'))
                cell.alignment = left_aligned
                cell.border = border
                
                # Column C: Description (Thai) - Placeholder
                cell = menu_sheet.cell(row=current_row, column=3, value="")
                cell.alignment = left_aligned
                cell.border = border
                
                # Column D: Price (direct value)
                price = item_dict.get('price', 0.0)
                cell = menu_sheet.cell(row=current_row, column=4, value=price)
                cell.number_format = '#,##0.00'; cell.alignment = center_aligned; cell.border = border
                
                # Column E: NET Price (calculated)
                net_price = round(price * 1.177)
                cell = menu_sheet.cell(row=current_row, column=5, value=net_price)
                cell.number_format = '#,##0'; cell.alignment = center_aligned; cell.border = border

            current_row += 1

    logging.info(f"'Menu Items' sheet created with linked cells. Last populated row: {current_row - 1}")
    return menu_sheet

def create_hungry_hub_proposal(json_file='./menu_bundles.json', menu_file='./categorized_menu_items.txt', output_file='HH_Proposal_Generated.xlsx'):
    """
    Creates the main Excel proposal file with three sheets:
    1. HH Proposal - the main proposal sheet
    2. Menu Items - categorized menu items (linked to Raw Menu Items)
    3. Raw Menu Items - flattened list of all menu items (editable)
    """
    # Dump the contents of the menu file for debugging
    dump_file_contents(menu_file)
    
    # Step 1: Read bundle data
    try:
        json_abs_path = os.path.abspath(json_file)
        logging.info(f"Attempting to read bundles from: {json_abs_path}")
        with open(json_abs_path, 'r', encoding='utf-8') as file:
            menu_bundles = json.load(file)
        logging.info(f"Successfully loaded {len(menu_bundles)} bundles from {json_file}")
        
        # Debug: Print loaded JSON to see the actual keys
        for i, bundle in enumerate(menu_bundles):
            logging.info(f"Bundle {i}: Keys = {list(bundle.keys())}")
            
        if len(menu_bundles) < 4:
            logging.warning(f"JSON file contains only {len(menu_bundles)} bundles, template expects 4.")
    except (FileNotFoundError, json.JSONDecodeError) as e:
        logging.error(f"Error reading JSON file '{json_file}': {e}")
        logging.warning("Using default menu bundles as fallback.")
        # Default bundles definition
        menu_bundles = [
            { "bundle_name": "Default A", "Suggested_bundle_price": "$3000", "Number_of_diners": 6, "category_portions": {"Category A": 3, "Category B": 4, "Category C": 2, "Category D": 1}, "Original_bundle_price": "$3500", "Discount_percentage": "15%", "Price_per_diner": "$500" },
            { "bundle_name": "Default B", "Suggested_bundle_price": "$4000", "Number_of_diners": 8, "category_portions": {"Category A": 4, "Category B": 6, "Category C": 3, "Category D": 2}, "Original_bundle_price": "$4800", "Discount_percentage": "20%", "Price_per_diner": "$500" },
            { "bundle_name": "Default C", "Suggested_bundle_price": "$5000", "Number_of_diners": 10, "category_portions": {"Category A": 5, "Category B": 8, "Category C": 4, "Category D": 3}, "Original_bundle_price": "$6000", "Discount_percentage": "17%", "Price_per_diner": "$500" },
            { "bundle_name": "Default D", "Suggested_bundle_price": "$6000", "Number_of_diners": 12, "category_portions": {"Category A": 6, "Category B": 10, "Category C": 5, "Category D": 4}, "Original_bundle_price": "$7200", "Discount_percentage": "17%", "Price_per_diner": "$500" }
        ]
        menu_bundles = menu_bundles[:4] # Ensure exactly 4

    # Extract all unique category keys from all bundles
    all_categories = set()
    for bundle in menu_bundles:
        if 'category_portions' in bundle and isinstance(bundle['category_portions'], dict):
            for category in bundle['category_portions'].keys():
                all_categories.add(category)
    
    # Sort categories alphabetically for consistent display
    all_categories = sorted(list(all_categories))
    logging.info(f"Found {len(all_categories)} unique categories in menu bundles: {all_categories}")

    # Parse menu items with our improved function
    menu_items_data = parse_menu_items_for_excel(menu_file)  # Changed from parse_menu_items to parse_menu_items_for_excel

    # Step 3: Create Workbook and 'HH Proposal' Sheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'HH Proposal'

    # Step 4: Define styles
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFFD4", end_color="FFFFD4", fill_type="solid")
    bright_yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Populate 'HH Proposal' Sheet ---
    # Header
    worksheet.merge_cells('A1:I1')
    header_cell = worksheet['A1']; header_cell.value = "Hungry Hub PARTY PACK PROPOSAL"
    header_cell.fill = red_fill; header_cell.font = white_font; header_cell.alignment = center_aligned
    worksheet.row_dimensions[1].height = 25
    # Restaurant Name
    worksheet['A2'] = "Restaurant Name:"; worksheet['A2'].font = bold_font
    worksheet.merge_cells('B2:I2'); worksheet['B2'] = restaurant_name.upper()
    worksheet['B2'].alignment = center_aligned; worksheet['B2'].font = Font(bold=True, color="FF0000")
    # Package Names Row 3
    worksheet['A3'] = "Package Name:"; worksheet['A3'].font = bold_font
    pack_definitions = [
        {"name": "Pack A", "columns": "B:C", "bundle_index": 0},
        {"name": "Pack B", "columns": "D:E", "bundle_index": 1},
        {"name": "Pack C", "columns": "F:G", "bundle_index": 2},
        {"name": "Pack D", "columns": "H:I", "bundle_index": 3}
    ]
    for pack in pack_definitions:
        if pack['bundle_index'] < len(menu_bundles):  # Make sure we don't go out of bounds
            cell_range = f"{pack['columns'].split(':')[0]}3:{pack['columns'].split(':')[1]}3"
            worksheet.merge_cells(cell_range)
            cell = worksheet[cell_range.split(':')[0]]
            # Get bundle name, defaulting to pack name if not found
            bundle_name = menu_bundles[pack['bundle_index']].get('bundle_name', pack['name'])
            cell.value = bundle_name; cell.alignment = center_aligned; cell.font = bold_font
        
    # HH Selling Price (NET) Row 4
    worksheet['A4'] = "HH Selling Price (NET)"; worksheet['A4'].font = bold_font
    for pack in pack_definitions:
        if pack['bundle_index'] < len(menu_bundles):  # Safety check
            col_start, col_end = pack['columns'].split(':')
            cell_range = f"{col_start}4:{col_end}4"; worksheet.merge_cells(cell_range)

            # Use 'Suggested_bundle_price' (exact match to JSON)
            price_value_raw = menu_bundles[pack['bundle_index']].get('Suggested_bundle_price', 0)
            price = 0.0
            try:
                if isinstance(price_value_raw, str):
                    price = float(price_value_raw.replace('$', '').replace(',', ''))
                elif isinstance(price_value_raw, (int, float)):
                    price = float(price_value_raw)
                else:
                    logging.warning(f"Unexpected type for Suggested_bundle_price: {type(price_value_raw)}. Using 0.")
                    price = 0.0
            except ValueError:
                logging.warning(f"Could not convert Suggested_bundle_price '{price_value_raw}' to float. Using 0.")
                price = 0.0

            cell = worksheet[col_start + '4']
            cell.value = price
            cell.number_format = '#,##0'; cell.alignment = center_aligned; cell.fill = light_yellow_fill

    # Max Diners Row 5
    worksheet['A5'] = "Max Diners / Set"; worksheet['A5'].font = bold_font
    for pack in pack_definitions:
        if pack['bundle_index'] < len(menu_bundles):  # Safety check
            col_start, col_end = pack['columns'].split(':')
            cell_range = f"{col_start}5:{col_end}5"; worksheet.merge_cells(cell_range)
            
            # Use 'Number_of_diners' (exact match to JSON)
            diners = menu_bundles[pack['bundle_index']].get('Number_of_diners', 0)
            cell = worksheet[col_start + '5']
            cell.value = diners; cell.number_format = '0'; cell.alignment = center_aligned; cell.fill = light_yellow_fill

    # Remarks Row 6
    worksheet['A6'] = "Remarks"; worksheet['A6'].font = bold_font
    for pack in pack_definitions:
        if pack['bundle_index'] < len(menu_bundles):  # Safety check
            col_start, col_end = pack['columns'].split(':')
            cell_range = f"{col_start}6:{col_end}6"; worksheet.merge_cells(cell_range)
            worksheet[col_start + '6'] = "1 Water / Person"; worksheet[col_start + '6'].alignment = center_aligned

    # Menu Section Header Row 7
    worksheet.merge_cells('A7:I7')
    menu_header = worksheet['A7']; menu_header.value = "Menu Section (portions from each section) - See Menu In Next Sheet"
    menu_header.fill = red_fill; menu_header.font = white_font; menu_header.alignment = center_aligned
    worksheet.row_dimensions[7].height = 25

    # DYNAMIC: Category Rows - starting from row 8
    current_row = 8
    for category_band in all_categories:
        worksheet[f'A{current_row}'] = category_band
        worksheet[f'A{current_row}'].font = bold_font
        
        logging.info(f"Processing {category_band} on row {current_row}")
        
        for pack in pack_definitions:
            if pack['bundle_index'] < len(menu_bundles):  # Safety check
                col_start, col_end = pack['columns'].split(':')
                cell_range = f"{col_start}{current_row}:{col_end}{current_row}"
                worksheet.merge_cells(cell_range)
                
                bundle_data = menu_bundles[pack['bundle_index']]
                bundle_name = bundle_data.get('bundle_name', f"Bundle {pack['bundle_index'] + 1}")
                
                logging.info(f"  Processing bundle: {bundle_name}")
                
                portions = 0
                
                if 'category_portions' in bundle_data:
                    if isinstance(bundle_data['category_portions'], dict):
                        logging.info(f"    Available categories: {list(bundle_data['category_portions'].keys())}")
                        
                        if category_band in bundle_data['category_portions']:
                            portions_raw = bundle_data['category_portions'][category_band]
                            try:
                                portions = int(portions_raw)
                            except (ValueError, TypeError):
                                portions = 0
                                logging.warning(f"    Could not convert '{portions_raw}' to int for {category_band}")
                        else:
                            logging.warning(f"    {category_band} not found in category_portions")
                    else:
                        logging.warning(f"    category_portions is not a dict: {type(bundle_data['category_portions'])}")
                else:
                    logging.warning(f"    No category_portions key found in bundle")
                
                cell = worksheet[col_start + str(current_row)]
                cell.value = portions
                cell.number_format = '0'
                cell.alignment = center_aligned
                cell.fill = light_yellow_fill
                
                logging.info(f"    Set cell {col_start}{current_row} to {portions}")
        
        current_row += 1  # Move to next row for next category

    # Total Dishes Row - after all category rows
    total_dishes_row = current_row
    worksheet[f'A{total_dishes_row}'] = "Total Dishes"; worksheet[f'A{total_dishes_row}'].font = bold_font
    for pack in pack_definitions:
        if pack['bundle_index'] < len(menu_bundles):  # Safety check
            col_start, col_end = pack['columns'].split(':')
            cell_range = f"{col_start}{total_dishes_row}:{col_end}{total_dishes_row}"; worksheet.merge_cells(cell_range)
            total_dishes = 0; bundle_data = menu_bundles[pack['bundle_index']]
            if 'category_portions' in bundle_data and isinstance(bundle_data['category_portions'], dict):
                total_dishes = sum(bundle_data['category_portions'].values())
            cell = worksheet[f"{col_start}{total_dishes_row}"]
            cell.value = total_dishes; cell.number_format = '0'; cell.alignment = center_aligned

    # Avg Price Header Row - 1 row after Total Dishes
    avg_price_header_row = total_dishes_row + 1
    worksheet.merge_cells(f'A{avg_price_header_row}:I{avg_price_header_row}')
    price_header = worksheet[f'A{avg_price_header_row}']; price_header.value = "Average NET Selling Price / Discounts"
    price_header.fill = red_fill; price_header.font = white_font; price_header.alignment = center_aligned
    worksheet.row_dimensions[avg_price_header_row].height = 25

    # Avg NET Selling Price Row (Original Price) - 1 row after the header
    avg_price_row = avg_price_header_row + 1
    worksheet[f'A{avg_price_row}'] = "Average NET Selling Price"; worksheet[f'A{avg_price_row}'].font = bold_font
    for pack in pack_definitions:
        if pack['bundle_index'] < len(menu_bundles):  # Safety check
            col_start, col_end = pack['columns'].split(':')
            cell_range = f"{col_start}{avg_price_row}:{col_end}{avg_price_row}"; worksheet.merge_cells(cell_range)

            # Use 'Original_bundle_price' (exact match to JSON)
            price_value_raw = menu_bundles[pack['bundle_index']].get('Original_bundle_price', 0)
            price = 0.0
            try:
                if isinstance(price_value_raw, str):
                    price = float(price_value_raw.replace('$', '').replace(',', ''))
                elif isinstance(price_value_raw, (int, float)):
                    price = float(price_value_raw)
                else:
                    logging.warning(f"Unexpected type for Original_bundle_price: {type(price_value_raw)}. Using 0.")
                    price = 0.0
            except ValueError:
                logging.warning(f"Could not convert Original_bundle_price '{price_value_raw}' to float. Using 0.")
                price = 0.0

            cell = worksheet[col_start + str(avg_price_row)]
            cell.value = price
            cell.number_format = '#,##0'; cell.alignment = center_aligned

    # Average Discount Row - 1 row after Avg NET Selling Price
    discount_row = avg_price_row + 1
    worksheet[f'A{discount_row}'] = "Average Discount"; worksheet[f'A{discount_row}'].font = bold_font
    for pack in pack_definitions:
        if pack['bundle_index'] < len(menu_bundles):  # Safety check
            col_start, col_end = pack['columns'].split(':')
            cell_range = f"{col_start}{discount_row}:{col_end}{discount_row}"; worksheet.merge_cells(cell_range)
            
            # Use 'Discount_percentage' (exact match to JSON)
            discount_value_raw = menu_bundles[pack['bundle_index']].get('Discount_percentage', '0%')
            try:
                if isinstance(discount_value_raw, str):
                    discount_val = float(discount_value_raw.replace('%', '')) / 100.0
                    number_format = '0%'
                elif isinstance(discount_value_raw, (int, float)):
                    discount_val = float(discount_value_raw)
                    if discount_val > 1:
                        discount_val = discount_val / 100.0
                    number_format = '0%'
                else:
                    logging.warning(f"Unexpected type for Discount_percentage: {type(discount_value_raw)}. Using 0%.")
                    discount_val = 0.0
                    number_format = '0%'
            except ValueError:
                discount_val = discount_value_raw
                number_format = '@'
            cell = worksheet[col_start + str(discount_row)]
            cell.value = discount_val; cell.number_format = number_format; cell.alignment = center_aligned

    # Net Price Per Person Row - 1 row after Average Discount
    price_per_person_row = discount_row + 1
    worksheet[f'A{price_per_person_row}'] = "Net Price / Person"; worksheet[f'A{price_per_person_row}'].font = bold_font
    for pack in pack_definitions:
        if pack['bundle_index'] < len(menu_bundles):  # Safety check
            col_start, col_end = pack['columns'].split(':')
            cell_range = f"{col_start}{price_per_person_row}:{col_end}{price_per_person_row}"; worksheet.merge_cells(cell_range)

            # Use 'Price_per_diner' (exact match to JSON)
            price_per_diner_val = menu_bundles[pack['bundle_index']].get('Price_per_diner', '0')
            display_text = "0 / Person"
            try:
                price_num = 0.0
                if isinstance(price_per_diner_val, str):
                    price_num = float(price_per_diner_val.replace('$', '').replace(',', ''))
                elif isinstance(price_per_diner_val, (int, float)):
                    price_num = float(price_per_diner_val)
                else:
                    logging.warning(f"Unexpected type for Price_per_diner: {type(price_per_diner_val)}. Using 0.")

                display_text = f"{price_num:,.0f} / Person"

            except ValueError:
                logging.warning(f"Could not convert Price_per_diner '{price_per_diner_val}' to number. Displaying as is.")
                display_text = f"{price_per_diner_val} / Person"
            except Exception as e:
                logging.error(f"Error formatting Price_per_diner '{price_per_diner_val}': {e}. Displaying as is.")
                display_text = f"{price_per_diner_val} / Person"

            cell = worksheet[col_start + str(price_per_person_row)]
            cell.value = display_text
            cell.alignment = center_aligned; cell.fill = bright_yellow_fill; cell.font = bold_font
        
    # Adjustment Info Rows - 2 rows after Net Price Per Person
    adjustment_row = price_per_person_row + 1
    worksheet[f'A{adjustment_row}'] = "You can Adjust"; worksheet[f'A{adjustment_row}'].font = Font(italic=True)
    
    dont_adjust_row = adjustment_row + 1
    worksheet[f'A{dont_adjust_row}'] = "Don't Adjust (Formula)"; worksheet[f'A{dont_adjust_row}'].font = Font(italic=True)
    
    adjustable_rows = [4, 5]
    for row in adjustable_rows: # Fill adjustable rows
        for pack in pack_definitions:
            if pack['bundle_index'] < len(menu_bundles):  # Safety check
                col_start, col_end = pack['columns'].split(':'); worksheet[f"{col_start}{row}"].fill = light_yellow_fill
            
    # Apply borders to all cells - calculate max row based on our dynamic structure
    max_row = dont_adjust_row
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=9):
        for cell in row:
            is_merged = any(cell.coordinate in merged_range for merged_range in worksheet.merged_cells.ranges)
            is_top_left = any(cell.coordinate == merged_range.coord.split(':')[0] for merged_range in worksheet.merged_cells.ranges if cell.coordinate in merged_range)
            if not is_merged or is_top_left: cell.border = border
            
    # Set column widths
    worksheet.column_dimensions['A'].width = 25
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']: worksheet.column_dimensions[col_letter].width = 15

    # Create the 'Raw Menu Items' sheet FIRST (we need the row mappings for the Menu Items sheet)
    raw_menu_sheet, menu_item_row_map = create_raw_menu_sheet(workbook, menu_items_data)
    
    # Create the 'Menu Items' sheet with references to the Raw Menu Items sheet
    create_menu_sheet(workbook, menu_items_data, menu_item_row_map)
    
    # Move the Raw Menu Items sheet to the end for better organization
    workbook.move_sheet(raw_menu_sheet, -1)  # Move to the end

    # Get all sheet names in current order
    sheet_names = workbook.sheetnames

    # Find the index of 'HH Proposal'
    hh_proposal_index = sheet_names.index('HH Proposal')

    # If 'HH Proposal' is not already first, move it to be first
    if hh_proposal_index != 0:
        hh_proposal_sheet = workbook['HH Proposal']
        workbook._sheets.remove(hh_proposal_sheet)
        workbook._sheets.insert(0, hh_proposal_sheet)

    # Make the 'HH Proposal' sheet active
    workbook.active = workbook['HH Proposal']
    
    # Save the workbook
    try:
        output_abs_path = os.path.abspath(output_file)
        logging.info(f"Attempting to save workbook to: {output_abs_path}")
        workbook.save(output_abs_path)
        return output_abs_path
    except PermissionError:
        logging.error(f"Permission denied. Could not save '{output_file}'. Check if the file is open or permissions.")
        return f"Error: Permission denied saving '{output_file}'."
    except Exception as e:
        logging.error(f"Error saving Excel file '{output_file}': {e}", exc_info=True)
        return None

@app.route('/api/upload', methods=['POST'])
def upload_images():
    # Get restaurant name from form data
    restaurant_name = request.form.get("restaurantName")
    
    if not restaurant_name:
        return jsonify({"message": "Restaurant name is required"}), 400

    # Create a directory for the restaurant (use secure_filename to sanitize input)
    restaurant_folder = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(restaurant_name))
    if not os.path.exists(restaurant_folder):
        os.makedirs(restaurant_folder)

    # Get uploaded files
    files = request.files.getlist("images")
    
    if not files:
        return jsonify({"message": "No images uploaded"}), 400

    saved_files = []
    
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(restaurant_folder, filename)
            file.save(filepath)
            saved_files.append(filename)

    return jsonify({
        "message": f"{len(saved_files)} images uploaded successfully for '{restaurant_name}'!",
        "files": saved_files,
        "restaurant": restaurant_name,
        "folder_path": restaurant_folder,
    }), 200

@app.route('/api/process-menu', methods=['POST'])
def process_menu():
    try:
        # Get restaurant name from request
        restaurant_name = request.json.get('restaurantName')
        num_categories = request.json.get('numCategories', 4)  # Default to 4
        if not restaurant_name:
            return jsonify({"error": "Restaurant name is required"}), 400

        # Get API key from headers
        api_key = request.headers.get('X-API-Key', gemini_api_key)

        # Construct path to uploaded images
        restaurant_folder = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(restaurant_name))

        # Check if folder exists
        if not os.path.exists(restaurant_folder):
            return jsonify({
                "error": f"No images found for restaurant '{restaurant_name}'. Please upload images first."
            }), 404

        # Get all image files in the folder
        image_paths = [
            os.path.join(restaurant_folder, f)
            for f in os.listdir(restaurant_folder)
            if allowed_file(f)
        ]

        if not image_paths:
            return jsonify({
                "error": f"No valid images found in the folder for restaurant '{restaurant_name}'. Please upload menu images first."
            }), 404

        # Process with dynamic category count
        price_categories, confidence_scores = process_menu_images(
            api_key, 
            restaurant_folder, 
            num_price_categories=num_categories
        )

        if price_categories:
            # --- Output Results ---
            logging.info("\n========== FINAL MENU ITEMS BY PRICE CATEGORY ==========")
            # Print to console
            for category_letter, items in sorted(price_categories.items()):
                if items:
                    # Sort items within the category by price (desc) for printing
                    items_sorted = sorted(items, key=lambda x: x['price'], reverse=True)
                    category_min_price = items_sorted[-1]['price']
                    category_max_price = items_sorted[0]['price']
                    print(f"\n----- Category {category_letter} (${category_min_price:.2f} - ${category_max_price:.2f}) -----")
                    print(f"{len(items)} items:")

                    for item_dict in items_sorted:
                        # Display Name, Price, and the SEMANTIC Category extracted by Claude
                        print(f"  ${item_dict['price']:.2f} - {item_dict['name']} ({item_dict['category']})") # Include semantic category in print

            # Save categorized items to a file
            try:
                output_path = os.path.join(
                    app.config['UPLOAD_FOLDER'],
                    secure_filename(restaurant_name),
                    f"categorized_menu_items_{secure_filename(restaurant_name)}.txt"
                )
                logging.info(f"\nSaving results to: {output_path}")
                with open(output_path, "w", encoding='utf-8') as f:
                    f.write("===== MENU ITEMS BY PRICE CATEGORY (Generated by Script) =====\n")
                    f.write(f"Processed images from: {os.path.abspath(restaurant_folder)}\n")
                    # Optionally add confidence score info here if desired
                    # avg_conf = (sum(confidence_scores.values()) / len(confidence_scores)) * 100 if confidence_scores else 0
                    # f.write(f"Average Extraction Confidence: {avg_conf:.1f}%\n")

                    for category_letter, items in sorted(price_categories.items()):
                        if items:
                            # Sort items within the category by price (desc) for writing
                            items_sorted = sorted(items, key=lambda x: x['price'], reverse=True)
                            category_min_price = items_sorted[-1]['price']
                            category_max_price = items_sorted[0]['price']
                            f.write(f"\n----- Category {category_letter} (${category_min_price:.2f} - ${category_max_price:.2f}) -----\n")

                            for item_dict in items_sorted:
                                # Write in the format expected by the *other* script,
                                # BUT let's include the semantic category in parentheses.
                                # The other script's parser might need adjustment if it can't handle this.
                                f.write(f"  ${item_dict['price']:.2f} - {item_dict['name']} ({item_dict['category']})\n")
                logging.info(f"Results successfully saved to {output_path}")
            except IOError as e:
                logging.error(f"Error saving results to file '{output_path}': {e}")
            except Exception as e:
                logging.error(f"An unexpected error occurred during file writing: {e}", exc_info=True)
        else:
            logging.error("Processing finished, but no menu items were found or categorized.")

        return jsonify({
            "categories": price_categories,
            "confidence_scores": confidence_scores
        })

    except Exception as e:
        app.logger.error(f"Processing failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": "Menu processing failed",
            "error": str(e)
        }), 500

@app.route('/api/get-categorized-menu', methods=['POST'])
def get_categorized_menu_text():
    try:
        restaurant_name = request.json.get('restaurantName')
        if not restaurant_name:
            return jsonify({"error": "Restaurant name is required"}), 400

        file_path = os.path.join(
            app.config['UPLOAD_FOLDER'],
            secure_filename(restaurant_name),
            f"categorized_menu_items_{secure_filename(restaurant_name)}.txt"
        )

        if not os.path.exists(file_path):
            app.logger.error(f"Categorized menu file not found at {file_path}")
            return jsonify({"error": f"Categorized menu file not found for {restaurant_name}. It might not have been generated yet or an error occurred."}), 404

        with open(file_path, 'r', encoding='utf-8') as f:
            menu_text_content = f.read()
        
        return jsonify({"menuText": menu_text_content}), 200

    except Exception as e:
        app.logger.error(f"Error fetching categorized menu text: {str(e)}")
        return jsonify({"error": "Internal server error while fetching menu text", "details": str(e)}), 500

@app.route('/api/update-categorized-menu', methods=['POST'])
def update_categorized_menu_text():
    try:
        restaurant_name = request.json.get('restaurantName')
        updated_menu_text = request.json.get('menuText')
        
        if not restaurant_name:
            return jsonify({"error": "Restaurant name is required"}), 400
        if updated_menu_text is None:
            return jsonify({"error": "Menu text is required"}), 400

        file_path = os.path.join(
            app.config['UPLOAD_FOLDER'],
            secure_filename(restaurant_name),
            f"categorized_menu_items_{secure_filename(restaurant_name)}.txt"
        )

        # Create directory if it doesn't exist (shouldn't happen but just in case)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(updated_menu_text)
        
        app.logger.info(f"Updated menu text saved for restaurant: {restaurant_name}")
        return jsonify({"success": True, "message": "Menu text updated successfully"}), 200

    except Exception as e:
        app.logger.error(f"Error updating categorized menu text: {str(e)}")
        return jsonify({"error": "Internal server error while updating menu text", "details": str(e)}), 500

@app.route('/api/generate-bundles', methods=['POST'])
def generate_bundles():
    restaurant_name = request.json.get('restaurantName')
    average_spend = request.json.get('averagePrice')
    try:
        generate(restaurant_name, average_spend)
        # Check if output file exists
        menu_bundles_file = os.path.join(
            app.config['UPLOAD_FOLDER'],
            secure_filename(restaurant_name),
            f"menu_bundles_{secure_filename(restaurant_name)}.json"
        )
        if os.path.exists(menu_bundles_file):
            with open(menu_bundles_file, "r", encoding="utf-8") as f:
                bundles = json.load(f)
            return jsonify({"status": "success", "bundles": bundles}), 200
        else:
            return jsonify({"status": "error", "message": "menu_bundles.json not found"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/generate-proposal', methods=['POST'])
def generate_proposal():
    try:
        restaurant_name = request.json.get('restaurantName')
        if not restaurant_name:
            return jsonify({"error": "Restaurant name is required"}), 400

        # Construct paths
        categorized_menu_items_file = os.path.join(
            app.config['UPLOAD_FOLDER'],
            secure_filename(restaurant_name),
            f"categorized_menu_items_{secure_filename(restaurant_name)}.txt"
        )
        menu_bundle_json_file = os.path.join(
            app.config['UPLOAD_FOLDER'],
            secure_filename(restaurant_name),
            f"menu_bundles_{secure_filename(restaurant_name)}.json"
        )

        output_file = os.path.join(
            app.config['UPLOAD_FOLDER'],
            secure_filename(restaurant_name),
            f"HH_proposal_{secure_filename(restaurant_name)}_generated.xlsx"
        )

        # Check if categorized & menu bundle files exist
        if not os.path.exists(categorized_menu_items_file):
            return jsonify({"error": f"Categorized menu items file not found: {categorized_menu_items_file}"}), 404
        elif not os.path.exists(menu_bundle_json_file):
            return jsonify({"error": f"Menu bundles json file not found: {menu_bundle_json_file}"}), 404

        # Generate Excel file
        excel_file = create_hungry_hub_proposal(restaurant_name, menu_bundle_json_file, categorized_menu_items_file, output_file)

        # Validate the returned path
        if not excel_file or not os.path.exists(excel_file):
            return jsonify({"error": "Failed to generate Excel file"}), 500

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'HH_proposal_{secure_filename(restaurant_name)}_generated.xlsx'
        )

    except Exception as e:
        app.logger.error(f"API Error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)